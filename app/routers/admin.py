from fastapi import APIRouter, Depends, UploadFile, File, HTTPException, Header
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from app.database import get_db
from app.models import Song, User, Score
from app.schemas import SongCreate, SongUpdate
from app.utils import normalize_title
from typing import List
import openpyxl
import io
import os
import re
import requests
from bs4 import BeautifulSoup

ZASA_URL = "https://zasa.sakura.ne.jp/dp/run.php"
TITLE_FIXES: dict[str, str] = {
    "Muzik LoverZ": "Musik LoverZ",
    "Voo Boo Bamboleo": "Voo Doo Bamboleo",
    "ƒƒƒƒƒ": "fffff",
    "POLꓘAMAИIA": "POLꞰAMAИIA",
}
DIFFICULTY_MAP = {"5": "HYPER", "7": "ANOTHER", "9": "LEGGENDARIA"}
CELL_RE = re.compile(r"☆(\d+)\s*\(([0-9.]+)\)")
LINK_RE = re.compile(r"music\.php\?id=(\d{5})-([579])-[01]")


def _parse_zasa_cell(cell):
    a = cell.find("a")
    if not a:
        return None
    link_match = LINK_RE.search(a.get("href", ""))
    if not link_match:
        return None
    chart = DIFFICULTY_MAP.get(link_match.group(2))
    if not chart:
        return None
    cell_match = CELL_RE.search(a.get_text())
    if not cell_match:
        return None
    return {
        "zasa_id": link_match.group(1),
        "chart": chart,
        "level": int(cell_match.group(1)),
        "unofficial_level": float(cell_match.group(2)),
    }


def _fetch_zasa_songs():
    resp = requests.get(ZASA_URL, timeout=30)
    resp.encoding = resp.apparent_encoding
    soup = BeautifulSoup(resp.text, "html.parser")
    songs = []
    for row in soup.find_all("tr"):
        cells = row.find_all("td")
        if len(cells) < 4:
            continue
        title = cells[-1].get_text(strip=True)
        if not title:
            continue
        title = TITLE_FIXES.get(title, title)
        for cell in cells[:3]:
            info = _parse_zasa_cell(cell)
            if info:
                songs.append({**info, "title": title})
    return songs


# 관리자 인증
def verify_admin(x_admin_key: str = Header(...)):
    password = os.environ.get("ADMIN_PASSWORD")
    if not password or x_admin_key != password:
        raise HTTPException(status_code=401, detail="Unauthorized")

router = APIRouter(dependencies=[Depends(verify_admin)])


# 곡 목록
@router.get("/admin/songs")
def get_songs(db: Session = Depends(get_db)):
    return db.query(Song).order_by(Song.level, Song.chart, Song.title).all()


# 곡 단건 수정
@router.put("/admin/songs/{song_id}")
def update_song(song_id: int, song: SongUpdate, db: Session = Depends(get_db)):
    existing = db.query(Song).filter(Song.id == song_id).first()
    if not existing:
        raise HTTPException(status_code=404, detail="곡을 찾을 수 없습니다")
    existing.title = song.title
    existing.title_normalized = normalize_title(song.title)
    existing.level = song.level
    existing.chart = song.chart
    existing.unofficial_level = song.unofficial_level
    db.commit()
    db.refresh(existing)
    return {
        "id": existing.id,
        "title": existing.title,
        "chart": existing.chart,
        "level": existing.level,
        "unofficial_level": existing.unofficial_level,
        "zasa_id": existing.zasa_id,
    }


# 곡 삭제
@router.delete("/admin/songs/{song_id}")
def delete_song(song_id: int, db: Session = Depends(get_db)):
    song = db.query(Song).filter(Song.id == song_id).first()
    if not song:
        raise HTTPException(status_code=404, detail="곡을 찾을 수 없습니다")
    db.delete(song)
    db.commit()
    return {"message": "삭제 완료"}


# Excel 내보내기
@router.get("/admin/songs/export")
def export_songs(db: Session = Depends(get_db)):
    songs = db.query(Song).order_by(Song.level, Song.chart, Song.title).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "songs"
    ws.append(["id", "title", "chart", "level", "unofficial_level", "zasa_id"])
    for s in songs:
        ws.append([s.id, s.title, s.chart, s.level, s.unofficial_level, s.zasa_id])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=songs.xlsx"},
    )


# Excel 가져오기 (있으면 수정, 없으면 신규 추가)
@router.post("/admin/songs/import")
def import_songs(file: UploadFile = File(...), db: Session = Depends(get_db)):
    contents = file.file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active

    updated = 0
    added = 0
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        song_id = row[0] if row[0] else None
        title = row[1] if len(row) > 1 else None
        chart = row[2] if len(row) > 2 else None
        level = row[3] if len(row) > 3 else None
        unofficial_level = row[4] if len(row) > 4 else None
        zasa_id = row[5] if len(row) > 5 else None

        if not title or not chart or not level:
            skipped += 1
            continue

        song = None
        if song_id:
            song = db.query(Song).filter(Song.id == int(song_id)).first()

        if song:
            song.title = str(title)
            song.title_normalized = normalize_title(str(title))
            song.level = int(level)
            song.unofficial_level = float(unofficial_level) if unofficial_level is not None else None
            if zasa_id is not None:
                song.zasa_id = str(zasa_id) if zasa_id else None
            updated += 1
        else:
            existing = db.query(Song).filter(
                Song.title_normalized == normalize_title(str(title)),
                Song.chart == str(chart)
            ).first()
            if existing:
                skipped += 1
                continue
            t = str(title)
            db.add(Song(
                title=t,
                title_normalized=normalize_title(t),
                level=int(level),
                chart=str(chart),
                unofficial_level=float(unofficial_level) if unofficial_level is not None else None,
                zasa_id=str(zasa_id) if zasa_id else None,
            ))
            added += 1

    db.commit()
    return {"updated": updated, "added": added, "skipped": skipped}


# zasa 동기화
@router.post("/admin/songs/fetch-zasa")
def sync_zasa(db: Session = Depends(get_db)):
    try:
        songs = _fetch_zasa_songs()
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"zasa 크롤링 실패: {e}")

    updated = 0
    added = 0

    for s in songs:
        title_normalized = normalize_title(s["title"])
        existing = db.query(Song).filter(
            Song.title_normalized == title_normalized,
            Song.chart == s["chart"],
            Song.level == s["level"]
        ).first()
        if existing:
            if existing.unofficial_level != s["unofficial_level"] or existing.zasa_id != s["zasa_id"]:
                existing.unofficial_level = s["unofficial_level"]
                existing.zasa_id = s["zasa_id"]
                updated += 1
        else:
            db.add(Song(
                title=s["title"],
                title_normalized=title_normalized,
                level=s["level"],
                chart=s["chart"],
                unofficial_level=s["unofficial_level"],
                zasa_id=s["zasa_id"],
            ))
            added += 1

    db.commit()
    return {"updated": updated, "added": added}


# 유저 목록
@router.get("/admin/users")
def get_users(db: Session = Depends(get_db)):
    users = db.query(User).order_by(User.dj_name).all()
    return [
        {
            "id": u.id,
            "iidx_id": u.iidx_id,
            "dj_name": u.dj_name,
            "score_count": db.query(Score).filter(Score.user_id == u.id).count(),
        }
        for u in users
    ]


# 유저 삭제 (스코어 포함)
@router.delete("/admin/users/{user_id}")
def delete_user(user_id: int, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="유저를 찾을 수 없습니다")
    db.query(Score).filter(Score.user_id == user_id).delete()
    db.delete(user)
    db.commit()
    return {"message": "삭제 완료"}